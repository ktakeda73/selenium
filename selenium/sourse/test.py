import os
import sys
from selenium import webdriver
import openpyxl
from tkinter import filedialog
from selenium.webdriver.chrome.options import Options
from .test2 import *

def scraping_selenium():
    options = Options()
    options.add_argument('--headless')
    options.add_argument('--hide-scrollbars')

    typ = [('Excelファイル','*.xlsx')]

    filepath = filedialog.askopenfilename(filetypes=typ)
    v_wb = openpyxl.load_workbook(filepath)
    v_ws = v_wb.active

    print(v_ws.max_row+1)
    for i in range(2,v_ws.max_row+1):
        v_Excelfrom = v_ws['b'+str(i)].value
        v_Excelto = v_ws['c'+str(i)].value

        v_browser = webdriver.Chrome('./driver/chromedriver.exe',options=options)

        v_browser.get('http://transit.yahoo.co.jp')

        v_from = v_browser.find_element_by_name("from")
        v_to = v_browser.find_element_by_name('to')

        v_from.send_keys(v_Excelfrom)
        v_to.send_keys(v_Excelto)

        get_screenshot('検索画面',v_browser)

        v_to.submit()

        v_fare=v_browser.find_element_by_class_name('fare').text.replace('円','')
        v_fare_int=v_fare.replace(',','')

        print(v_fare_int)
        v_ws['e'+str(i)].value = int(v_fare_int)

        get_screenshot('結果画面',v_browser)

    v_wb.save(filepath)

    

